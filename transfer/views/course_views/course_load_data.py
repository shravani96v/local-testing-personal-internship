from transfer.models.course_model import Course
from openpyxl import load_workbook
from openpyxl import Workbook

def course_load_data(request, wb_obj):
    major_name_list = wb_obj.sheetnames
    subject_number_list = []
    title_list = []

    for major_name in major_name_list:
        ws_obj = wb_obj[major_name]
        for col in ws_obj.iter_cols(min_col = 2, max_col= 2 ):
            for cell in col:
                if cell.value not in subject_number_list:
                    subject_number_list.append(cell.value)
                    subject_number = cell.value

        for col in ws_obj.iter_cols(min_col = 3, max_col = 3):
            for cell in col:
                if cell.value not in title_list:
                    title_list.append(cell.value)
                    title=cell.value
                    #print(title)
    #print(len(title_list))
    #print(len(subject_number_list))


    #for idx in range(len(subject_number_list)-1):
    #    subject_number = subject_number_list[idx]
    #    title_text = title_list[idx]
    #    index = idx+1
    #    course_obj = Course(subject_number = subject_number, title = title_text)
    #    course_obj.save()

    #for number in subject_number_list:
    #    course_obj1 = Course(subject_number = number)
    #    course_obj1.save()

    #for title in title_list:
    #    course_obj2 = Course(title = title)
    #    course_obj2.save()

    #for title in title_list:
    #    course_obj = Course(title=title)
