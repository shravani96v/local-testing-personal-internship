from openpyxl import load_workbook
from transfer.models.major_model import Major
from openpyxl import Workbook


def major_load_data(request, wb_obj):
    #print(f'in import_major wb_obj is {type(wb_obj)}')
    #print(f'wb_obj is {dir(wb_obj)}')
    major_names = wb_obj.sheetnames
    #print(f'sheetnames are {major_names}')
    for idx,name in enumerate(major_names):
        major_data=Major(major_id = idx,major_name = name)
        major_data.save()
