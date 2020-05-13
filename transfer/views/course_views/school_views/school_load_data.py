from transfer.models.school_model import School
#from transfer.models.major_model import Major
from openpyxl import load_workbook
from openpyxl import Workbook


def school_load_data(request, wb_obj):
    major_name_list = wb_obj.sheetnames
    school_name_list = []

    for major_name in major_name_list:
        ws_obj = wb_obj[major_name]
        for col in ws_obj.iter_cols(min_col=1, max_col=1):
            for cell in col:
                if cell.value not in school_name_list:
                    school_name_list.append(cell.value)
                    school_name=cell.value
                    #print(school_name)
                    
    for idx,name in enumerate(school_name_list):
        school_obj = School(school_id = idx,school_name = name)#major_id,
        school_obj.save()
