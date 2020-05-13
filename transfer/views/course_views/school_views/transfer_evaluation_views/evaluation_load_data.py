from transfer.models.transfer_evaluation_model import TransferEvaluation
from openpyxl import load_workbook
from openpyxl import Workbook

def evaluation_load_data(request, wb_obj):
    major_name_list = wb_obj.sheetnames
    sem_year_taken = []
    status = []
    comment = []

    for major_name in major_name_list:
        ws_obj = wb_obj[major_name]
        for col in ws_obj.iter_cols(min_col = 4, max_col= 4 ):
            for cell in col:
                if cell.value not in sem_year_taken:
                    sem_year_taken.append(cell.value)
                    year = cell.value
                    print(year)

        for col in ws_obj.iter_cols(min_col = 5, max_col = 5):
            for cell in col:
                if cell.value not in status:
                    status.append(cell.value)
                    sta=cell.value
                    print(sta)

        for col in ws_obj.iter_cols(min_col = 9, max_col = 9):
            for cell in col:
                if cell.value not in comment:
                    comment.append(cell.value)
                    com=cell.value
                    print(com)
