from django.test import TestCase
from openpyxl import load_workbook

# Create your tests here

wb_obj = load_workbook(filename = '/home/unhmguest/Desktop/transfereval-docs/transfers_bymajor.xlsx', data_only=True)
print(wb_obj.sheetnames)
